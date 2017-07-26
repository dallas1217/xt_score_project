#coding=utf-8

import os
from enum import Enum
from openpyxl import Workbook
from openpyxl import load_workbook
from pos.position import *
from info.player_info import *

temp_excel = 'xt_score_template.xlsx'
temp_sheet = 'Sheet1'
res_excel = 'xt_score_test.xlsx'

wb = load_workbook(temp_excel)
ws = wb.get_sheet_by_name(temp_sheet)
res_path = os.path.abspath(res_excel)
    
if __name__ == '__main__':
    start_line = 33        #打分表开始位置
    stats_start_line = 34  #填充数据开始位置
    liver_stats = main()    #player_info的main函数
    print("正在生成数据，请稍等")
    
    for info in liver_stats:
        pos = info['position']
        player_name = info['name']
        stats = info['stats']

        if pos in POSITION.keys():
            pos_id = POSITION[pos]    #获取位置IP，用于修改打分卡内位置和球员名称
            for i in range(4,28,3):
                default_pos = ws['A%d' %i].value
                if pos_id == default_pos:
                    for row in ws.iter_rows(min_row = i,max_row = (i+2)):   # 模板分为三行，这里i的值到i+2即为要使用的模板列
                        for cell in row:
                            new_cell = ws['%s%d' %(cell.column,start_line)]
                            new_cell.value = cell.value
                            new_cell.style = cell.style
                            
                        if ws['A%d' %start_line].value == pos_id:         #A列位置如果和pos_id相等，修改球员名字和位置名字
                            ws['A%d' %start_line].value = player_name
                            ws['B%d' %start_line].value = pos
                            
                        start_line+=1
                        
            ### 根据位置进行填数           
            if pos == 'GK':
                for name in GK_STATS.keys():
                    ws['%s%d' % (GK_STATS[name],stats_start_line)].value = stats.get(name)
                print('生成[{0}位置，{1}球员]数据完毕'.format(pos,player_name)) 
            else:
                for name in OUTFIELD_STATS.keys():
                    ws['%s%d' % (OUTFIELD_STATS[name],stats_start_line)].value = stats.get(name)
                print('生成[{0}位置，{1}球员]数据完毕' .format(pos,player_name))
            stats_start_line += 3
                 
                    
    #保存表格
    wb.save(res_excel)
    print('生成数据完毕，结果表格路径为：',res_path)

